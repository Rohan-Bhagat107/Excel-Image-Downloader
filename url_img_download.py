# Input-> Directory (.xlsx)
# Output-> Downloaded images
# Type-> This script takes the directory path, column name having urls from user and download the images from the provided columns in folder(one folder for each excel file)

import os
import requests
import pandas as pd
import warnings
import time
from urllib.parse import urlparse
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook

warnings.filterwarnings("ignore", category=UserWarning, module='openpyxl')

IMAGE_EXTENSIONS = (".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp")

def validator(input_path):
    if os.path.exists(input_path) and os.path.isdir(input_path):
        print(f"Validation successful for {input_path}.")
        return True
    print(f"Error: Invalid folder path! {input_path}")
    return False
# Function for getting the unhidden columns from the excel
def get_unhidden_columns(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active
    unhidden_cols = []
    for col in ws.iter_cols(1, ws.max_column):
        col_letter = col[0].column_letter
        dim = ws.column_dimensions.get(col_letter)
        if not (dim and getattr(dim, 'hidden', False)):
            unhidden_cols.append(col[0].value)
    wb.close()
    return [col for col in unhidden_cols if col is not None]
#Function for detecting the unhidden columns having url(max 2)
def detect_url_columns(df, unhidden_columns, max_columns=2):
    url_scores = {}
    for col in df.columns:
        if col not in unhidden_columns:
            continue
        try:
            series = df[col].astype(str).dropna()
            count = series.str.startswith("http").sum()
            image_count = series.str.lower().str.endswith(IMAGE_EXTENSIONS).sum()
            score = count + image_count
            if score > 0:
                url_scores[col] = score
        except Exception:
            continue

    sorted_cols = sorted(url_scores.items(), key=lambda x: x[1], reverse=True)
    best_cols = [col for col, _ in sorted_cols[:max_columns]]

    if best_cols:
        print(f"Detected URL columns: {best_cols}")
    else:
        print("No URL columns detected.")

    return best_cols
# Function for reading the urls from the user provided column
def read_image_links(excel_file, user_columns=None):
    #Assuming that user has provided the unhidden columns if not then detecting the url column
    try:
        df = pd.read_excel(excel_file, engine='openpyxl')
        unhidden_columns = get_unhidden_columns(excel_file)
        print(f"Unhidden columns in '{os.path.basename(excel_file)}': {unhidden_columns}")
        if user_columns:
            url_columns = [col for col in user_columns if col in unhidden_columns]
            if not url_columns:
                print("None of the specified columns found or all are hidden.")
                return []
            print(f"Using user-specified columns: {url_columns}")
        else:
            url_columns = detect_url_columns(df, unhidden_columns, max_columns=2)
        # Checking weather excel has url column or not
        if not url_columns:
            return []

        image_links = []
        for col in url_columns:
            col_data = df[col].dropna().astype(str).str.strip()
            image_links.extend(col_data)

        return image_links

    except Exception as e:
        print(f"Error reading {os.path.basename(excel_file)}: {e}")
        return []
# Function for creating output folders in the output directory provided by user
def create_output_folder(base_name, image_count, output_dir='.'):
    folder_name = f"{base_name} -te-IN ({image_count})"
    folder_path = os.path.join(output_dir, folder_name)
    os.makedirs(folder_path, exist_ok=True)
    return folder_path
# Function for downloading the images  from url and saving them in output directory
def download_image(url, save_dir, retries=3):
    filename = os.path.basename(urlparse(url).path)
    if not filename:
        filename = f"image_{int(time.time())}.jpg"
    filepath = os.path.join(save_dir, filename)

    for attempt in range(1, retries + 1):
        try:
            print(f"Downloading (attempt {attempt}): {url}")
            response = requests.get(url, stream=True, timeout=30)
            response.raise_for_status()

            # Get expected content size
            content_length = response.headers.get('Content-Length')
            expected_size = int(content_length) if content_length and content_length.isdigit() else None

            # Write content to file in chunks
            temp_filepath = filepath + ".part"
            with open(temp_filepath, "wb") as file:
                total_written = 0
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        file.write(chunk)
                        total_written += len(chunk)

            # Verify file integrity
            if expected_size and total_written != expected_size:
                print(f"Incomplete download detected for '{filename}'. Expected {expected_size}, got {total_written}. Retrying...")
                os.remove(temp_filepath)
                raise ValueError("Incomplete download")
            else:
                os.rename(temp_filepath, filepath)
                print(f"Downloaded: {filename}")
                return

        except Exception as e:
            print(f"Attempt {attempt} failed for '{filename}': {e}")
            if os.path.exists(filepath + ".part"):
                os.remove(filepath + ".part")
            if attempt < retries:
                time.sleep(2 ** attempt)
            else:
                print(f"❌ Failed to download image '{filename}' after {retries} attempts.")

# Supporting functon for creating thread if any image takes time for downloading
def download_images(image_links, folder_path):
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = [executor.submit(download_image, url, folder_path) for url in image_links]
        for future in as_completed(futures):
            try:
                future.result()
            except Exception as e:
                print(f"Thread error: {e}")
# Function for processing the requirement
def process_excel_file(excel_file,output_dir, user_columns=None):
    base_name = os.path.splitext(os.path.basename(excel_file))[0]
    image_links = read_image_links(excel_file, user_columns=user_columns)

    print(f"\nProcessing: {base_name}")
    print(f"Total image links: {len(image_links)}")

    if not image_links:
        print("No valid image links found. Skipping.")
        return
    #Calling a function for creating the output folder.
    folder_path = create_output_folder(base_name, len(image_links), output_dir)
    download_images(image_links, folder_path)

if __name__ == '__main__':
    try:
        folder_path = input("Enter folder path containing Excel files: ").strip()
        columns_input = input("Enter column names containing URLs (comma-separated): ").strip()
        output_directory=input("Please enter folder path for storing the output: ")
        user_columns = [col.strip() for col in columns_input.split(",") if col.strip()] if columns_input else None
        # Calling validator function for validating user input.
        if validator(folder_path):
            if validator(output_directory):
                for root, dirs, files in os.walk(folder_path):
                    for file in files:
                        if "te-IN.xlsx" in file and file.lower().endswith((".xlsx", ".xls")):
                            full_path = os.path.join(root, file)
                            process_excel_file(full_path, output_directory, user_columns=user_columns)


    except Exception as e:
        print(f"Error: {e}")
